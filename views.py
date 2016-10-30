from django.shortcuts import render,redirect,render_to_response,get_object_or_404
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.contrib import messages
from django.contrib.auth import authenticate, login,logout 
from django.http import HttpResponse,HttpResponseRedirect
from web_MA_DB.models import Node, Category,Project,Usertype,Userfield,Organization, ProjectSheet, InvoiceSheet, Invoice, QuoteSheet
from django.contrib.auth.models import Permission, User
from web_MA_DB.forms import ProjectForm,SearchForm, YearSearchForm, InvoiceSearchForm, UserForm, UploadFileForm,ReconcilliationForm, StateSearchForm, FieldForm, TypeForm, IntExtForm, UpdateFieldForm,QuarterForm,CustomSearchForm,FinancialYearForm
# from chartit import PivotDataPool, PivotChart
from django.db.models import Sum, Q
import openpyxl
from openpyxl import Workbook
import MySQLdb
from datetime import date,datetime
import decimal
from openpyxl.formula import Tokenizer
import simplejson

STATES = ['VIC','QLD','NSW','SA''ACT','NZ','TAS','WA','NT','Overseas']
TYPE_CHOICES = [u.type_id for u in Usertype.objects.filter(~Q(type_id='NULL'))]
INFO="Invalid date input."

def index(request):
	node_list = Node.objects.order_by('node_id')
	context_dict = {'nodes':node_list}
	return render(request, 'web_MA_DB/index.html', context_dict)
def homepage(request):
	return render(request,'web_MA_DB/homepage.html',{})
def blank(request):
	return render(request,'web_MA_DB/blank.html',{})


# Display all projects in a specific node
def node(request, node_name_slug):
	context_dict={}

	try:
		context_dict['node_name_slug'] = node_name_slug
		node = Node.objects.get(slug=node_name_slug)
		context_dict['node_name'] = node.node_name

		projects = Project.objects.filter(node=node)
		context_dict['projects'] = projects

		context_dict['node'] = node

		form = CustomSearchForm()		
		context_dict['form'] = form
	except Node.DoesNotExist:
		pass

	return render(request,'web_MA_DB/node.html',context_dict)

def projects_list(request,node_name_slug):
	context_dict = {}
	try:
		context_dict['node_name_slug'] = node_name_slug
		node = Node.objects.get(slug=node_name_slug)
		context_dict['node_name'] = node.node_name

		projects = Project.objects.filter(node=node)
		context_dict['projects'] = projects

		context_dict['node'] = node
	except Node.DoesNotExist:
		pass

	return render(request,'web_MA_DB/projects_list.html',context_dict)

def cats(start,end,projects):
	cates = ['1','2','3']
	data=[]

	if projects.filter(pro_date__gte=start,pro_date__lte=end).exists():
		projects = projects.filter(pro_date__gte=start,pro_date__lte=end)
		
		for t in cates:
			data.append(t)
			if projects.filter(category=t).exists():
				data.append(list(projects.filter(category=t).aggregate(Sum('subtotal')).values())[0])
				data.append(list(projects.filter(category=t).aggregate(Sum('cus_count')).values())[0])
				data.append(projects.filter(category=t).count())
				data.append(list(projects.filter(category=t).aggregate(Sum('num_sample')).values())[0])
			else:
				data.append(0)
				data.append(0)
				data.append(0)
				data.append(0)

		data=[data[x:x+5] for x in range(0, len(data),5)]
	else:
		data = [[0,0,0,0,0]]

	return data


def types(start,end,projects):

	types = [ t.type_id for t in Usertype.objects.filter(~Q(type_id='NULL'))]
	data=[]

	if projects.filter(pro_date__gte=start,pro_date__lte=end).exists():
		projects = projects.filter(pro_date__gte=start,pro_date__lte=end)
		
		for t in types:
			data.append(t)
			if projects.filter(usertype=t).exists():
				data.append(list(projects.filter(usertype=t).aggregate(Sum('subtotal')).values())[0])
				data.append(list(projects.filter(usertype=t).aggregate(Sum('cus_count')).values())[0])
				data.append(projects.filter(usertype=t).count())
				data.append(list(projects.filter(usertype=t).aggregate(Sum('num_sample')).values())[0])
			else:
				data.append(0)
				data.append(0)
				data.append(0)
				data.append(0)

		data=[data[x:x+5] for x in range(0, len(data),5)]
	else:
		data = [[0,0,0,0,0]]

	return data


def fields(start,end,projects):

	fields = [ f.field_id for f in Userfield.objects.filter(~Q(field_id='NULL'))]
	data=[]

	if projects.filter(pro_date__gte=start,pro_date__lte=end).exists():
		projects = projects.filter(pro_date__gte=start,pro_date__lte=end)
		
		for t in fields:
			data.append(t)
			if projects.filter(userfield=t).exists():
				data.append(list(projects.filter(userfield=t).aggregate(Sum('subtotal')).values())[0])
				data.append(list(projects.filter(userfield=t).aggregate(Sum('cus_count')).values())[0])
				data.append(projects.filter(userfield=t).count())
				data.append(list(projects.filter(userfield=t).aggregate(Sum('num_sample')).values())[0])
			else:
				data.append(0)
				data.append(0)
				data.append(0)
				data.append(0)

		data=[data[x:x+5] for x in range(0, len(data),5)]
	
	else: 
		data = [[0,0,0,0,0]]

	return data


def states(start,end,projects):
	data = []
	if projects.filter(pro_date__gte=start,pro_date__lte=end).exists():
		projects = projects.filter(pro_date__gte=start,pro_date__lte=end)
			
		for s in STATES:
			data.append(s)			
			if projects.filter(state=s).exists():
				data.append(list(projects.filter(state=s).aggregate(Sum('subtotal')).values())[0])
				data.append(list(projects.filter(state=s).aggregate(Sum('cus_count')).values())[0])
				data.append(projects.filter(state=s).count())
				data.append(list(projects.filter(state=s).aggregate(Sum('num_sample')).values())[0])
			else:
				data.append(0)
				data.append(0)
				data.append(0)
				data.append(0)

		data = [data[x:x+5] for x in range(0, len(data),5)]
	else:
		data = [[0,0,0,0,0]]

	return data

def write_state_excel(filename, context_dict):
	filename= filename
	sheetname = 'sheet1'
	wb1 = openpyxl.Workbook()
	ws1 = wb1.create_sheet(index=0,title = sheetname)
	wb1.save(filename)

	wb2 =openpyxl.load_workbook(filename)
	ws = wb2.get_sheet_by_name(sheetname)

	if 'states' in context_dict.keys():
		ws.cell(row=1,column=1).value = 'State'
	if  'fields' in context_dict.keys():
		ws.cell(row=1,column=1).value = 'UserDefined_2'
	if 'types' in context_dict.keys():
		ws.cell(row=1,column=1).value = 'UserDefined_1'
	if 'intexts' in context_dict.keys():
		ws.cell(row=1,column=1).value = 'Int/Ext'
	if 'nodes' in context_dict.keys():
		ws.cell(row=1,column=1).value = 'Node'
	ws.cell(row=1,column=2).value = 'Sales($)'
	ws.cell(row=1,column=3).value = 'No. Customer'
	ws.cell(row=1,column=4).value = 'No.Invoice'

	row = 2
	col = 1

	for array in context_dict['data']:
		for i in range(len(array)):
			ws.cell(row=row,column=col+i).value = array[i]
		row+=1

	wb2.save(filename)

def write_to_excel(filename,data,name):
	filename= filename
	sheetname = 'sheet1'
	wb1 = openpyxl.Workbook()
	ws1 = wb1.create_sheet(index=0,title = sheetname)
	wb1.save(filename)

	wb2 =openpyxl.load_workbook(filename)
	ws = wb2.get_sheet_by_name(sheetname)

	if 'states' in name:
		ws.cell(row=1,column=1).value = 'State'
	if  'fields' in name:
		ws.cell(row=1,column=1).value = 'UserDefined_2'
	if 'types' in name:
		ws.cell(row=1,column=1).value = 'UserDefined_1'
	if 'intexts' in name:
		ws.cell(row=1,column=1).value = 'Int/Ext'
	if 'nodes' in name:
		ws.cell(row=1,column=1).value = 'Node'
	if 'cats' in name:
		ws.cell(row=1,column=1).value = 'Category'
	ws.cell(row=1,column=2).value = 'Sales($)'
	ws.cell(row=1,column=3).value = 'No. Customer'
	ws.cell(row=1,column=4).value = 'No.Invoice'
	ws.cell(row=1,column=5).value = 'No.Sample'

	row = 2
	col = 1

	for array in data:
		for i in range(len(array)):
			ws.cell(row=row,column=col+i).value = array[i]
		row+=1

	wb2.save(filename)


def write_sum_excel(filename,tb_title,data):
	filename= filename
	sheetname = 'sheet1'
	wb1 = openpyxl.Workbook()
	ws1 = wb1.create_sheet(index=0,title = sheetname)
	wb1.save(filename)

	wb2 =openpyxl.load_workbook(filename)
	ws = wb2.get_sheet_by_name(sheetname)

	ws.cell(row=1,column=1).value = tb_title
	ws.cell(row=2,column=1).value = 'Sales($)'
	ws.cell(row=2,column=2).value = 'No. Customer'
	ws.cell(row=2,column=3).value = 'No.Invoice'
	ws.cell(row=2,column=4).value = 'No.Sample'

	row = 3
	col = 1

	for array in data:
		for i in range(len(array)):
			ws.cell(row=row,column=col+i).value = array[i]
		row+=1

	wb2.save(filename)



def sum_all(start,end,projects):
	data = []
	if projects.filter(pro_date__gte=start,pro_date__lte=end).exists():
		projects = projects.filter(pro_date__gte=start,pro_date__lte=end)
		
		data.append(list(projects.aggregate(Sum('subtotal')).values())[0])
		data.append(list(projects.aggregate(Sum('cus_count')).values())[0])
		data.append(projects.count())
		data.append(list(projects.aggregate(Sum('num_sample')).values())[0])

		data = [data[x:x+4] for x in range(0, len(data),4)]
	else:
		data = [[0,0,0,0]]

	return data


def in_node_results(request,node_name_slug):
	context_dict={}
	node = Node.objects.get(slug=node_name_slug)
	context_dict['node'] = node
	form = CustomSearchForm(request.GET)
	projects = Project.objects.filter(node=node)

	if form.is_valid():
		paras= form.cleaned_data		
		start = paras['from_date']
		end = paras['to_date']
		if 'Category' in paras['choices']:
			context_dict['Category'] = cats(start,end,projects)
		if 'UserDefined_1' in paras['choices']:
			context_dict['UserDefined_1'] = types(start,end,projects)
		if 'UserDefined_2' in paras['choices']:
			context_dict['UserDefined_2'] = fields(start,end,projects) 
		if 'State' in paras['choices']:
			context_dict['State'] = states(start,end,projects)
		if 'Sum' in paras['choices']:
			context_dict['Sum'] = sum_all(start,end,projects)
		context_dict['from'] = start
		context_dict['to'] = end


		if 'state_excel' in request.POST:	
			all_states = STATES			
			filename = node.node_id + "_byState_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			xlsx_data = write_to_excel(filename, context_dict['State'],{'states':all_states})
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response

		if 'type_excel' in request.POST:	
			all_types =[ t.type_id for t in Usertype.objects.filter(~Q(type_id='NULL'))]	
			filename = node.node_id +"_byUserDefined1_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			xlsx_data = write_to_excel(filename, context_dict['UserDefined_1'],{'types':types})
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response
		if 'field_excel' in request.POST:
			all_fields = [ f.field_id for f in Userfield.objects.filter(~Q(field_id='NULL'))]
			filename = node.node_id +"_byUserDefined2_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			xlsx_data = write_to_excel(filename, context_dict['UserDefined_2'],{'fields':all_fields})
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response
		if 'cat_excel' in request.POST:
			all_cats = ['1','2','3']
			filename = node.node_id +"_byCategory_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			xlsx_data = write_to_excel(filename, context_dict['Category'],{'cats':all_cats})
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response
		
		if 'sum_excel' in request.POST:
			filename = node.node_id +"_CustomerDetail_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			tb_title = node.node_id+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)
			xlsx_data = write_sum_excel(filename,tb_title ,context_dict['Sum'])
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response

	else:
		context_dict['info'] ='Invalid date or choices input.'
	return render(request,'web_MA_DB/in_node_results.html',context_dict)


def project_detail(request,pro_id):
	context_dict={}
	
	try:
		p = Project.objects.get(pro_id=pro_id)
	except:
		p = None
	context_dict['pro_id']=p.pro_id
	context_dict['p']=p

	return render(request,'web_MA_DB/project_detail.html',context_dict)



def add_project(request,node_name_slug):
	context_dict={}
	try:
		node = Node.objects.get(slug=node_name_slug)
	except Node.DoesNotExist:
		node = None
	context_dict['node']=node
	if request.method == 'POST':
		form = ProjectForm(request.POST)
		if form.is_valid():
			data = form.cleaned_data
			if node:
				# insert a project				
				proj_cat = Category.objects.get(cat_id=data['category'])
				proj_type = Usertype.objects.get(type_id=data['usertype'])
				proj_field = Userfield.objects.get(field_id=data['userfield'])
				p = Project.objects.create(node=node,organization =data['organization'],
					usertype = proj_type,userfield = proj_field,person = data['person'],pro_date=data['pro_date'],
					num_sample=data['num_sample'],category=proj_cat,description=data['description'],service=data['service'],instrument=data['instrument'],
					int_ext=data['int_ext'],state=data['state'],country=data['country'],subtotal=data['subtotal'],
					cus_count = data['cus_count'])
				p.save()
				
				
				return HttpResponseRedirect(reverse('node',args=[node.slug]))
		else:
			print (form.errors)
	else:
		form = ProjectForm()
	context_dict['form'] = form
	return render(request,'web_MA_DB/add_project.html',context_dict)

		
def client_volume(request):
	context_dict={}
	if request.method == 'POST':
		form = QuarterForm(request.POST)
		if form.is_valid():
			start = form.cleaned_data['from_date']
			end = form.cleaned_data['to_date']
			
			if Project.objects.filter(Q(pro_date__gte=start), Q(pro_date__lte=end)).exists():
				projects = Project.objects.filter(Q(pro_date__gte=start),Q(pro_date__lte=end))
				
				# context_dict['start'] = start
				# context_dict['end'] = end

				context_dict['AWRI']= 0
				context_dict['Murdoch']= 0
				context_dict['UQ']=0
				context_dict['UM']=0
				context_dict['UWA']=0
	
				if projects.filter(node='AWRI'):
					context_dict['AWRI']+=list(projects.filter(node='AWRI').aggregate(Sum('cus_count')).values())[0]

				if projects.filter(node='Murdoch'):
					context_dict['Murdoch']+=list(projects.filter(node='Murdoch').aggregate(Sum('cus_count')).values())[0]

				if projects.filter(node='UQ'):
					context_dict['UQ']+=list(projects.filter(node='UQ').aggregate(Sum('cus_count')).values())[0]
	
				if projects.filter(node='UM'):
					context_dict['UM']+=list(projects.filter(node='UM').aggregate(Sum('cus_count')).values())[0]
	
				if projects.filter(node='UWA'):
					context_dict['UWA']+=list(projects.filter(node='UWA').aggregate(Sum('cus_count')).values())[0]

				
				return render(request,'web_MA_DB/client_volume.html',context_dict)
			else:
				return render(request,'web_MA_DB/client_volume.html',{"info":INFO})

	else:
		form = QuarterForm()
		return render(request,'web_MA_DB/client_volume.html',{'form':form})


def node_chart(request,node_name_slug):
	def get_aggregation(node_name_slug):
		res = {}
		res.setdefault('1',0)
		res.setdefault('2',0)
		res.setdefault('3',0)
		nd= Node.objects.filter(slug=node_name_slug)
		projects = Project.objects.filter(node=nd)

		if projects.filter(category='1'):
			res['1']+=list(projects.filter(category='1').aggregate(Sum('subtotal')).values())[0]
		if projects.filter(category='2') :
			res['2']+=list(projects.filter(category='2').aggregate(Sum('subtotal')).values())[0]
		if projects.filter(category='3') :
			res['3']+=list(projects.filter(category='3').aggregate(Sum('subtotal')).values())[0]
		return res

	
	data = get_aggregation(node_name_slug)

	return render(request,'web_MA_DB/node_chart.html',{'cat1':data['1'],'cat2':data['2'],'cat3':data['3']})


def fund_chart(request):
	
	if request.method == 'POST':
		form = QuarterForm(request.POST)
		if form.is_valid():
			start = form.cleaned_data['from_date']
			end = form.cleaned_data['to_date']
			if Project.objects.filter(Q(pro_date__gte=start), Q(pro_date__lte=end)).exists():
				projects = Project.objects.filter(Q(pro_date__gte=start),Q(pro_date__lte=end))
			
				data ={'1':0,'2':0,'3':0}
				if projects.filter(category='1'):
					data['1']+=list(projects.filter(category='1').aggregate(Sum('subtotal')).values())[0]
				if projects.filter(category='2'):
					data['2']+=list(projects.filter(category='2').aggregate(Sum('subtotal')).values())[0]
				if projects.filter(category='3'):
					data['3']+=list(projects.filter(category='3').aggregate(Sum('subtotal')).values())[0]
				return render(request,'web_MA_DB/fund_chart.html',{'cat1':data['1'],'cat2':data['2'],'cat3':data['3']})

			else:
				return render(request,'web_MA_DB/fund_chart.html',{'info':INFO})
	else:
		form = QuarterForm()
		return render(request,'web_MA_DB/fund_chart.html',{'form':form})


def invoice_number(request):
	if request.method =='POST':
		form = QuarterForm(request.POST)
		if form.is_valid():
			start = form.cleaned_data['from_date']
			end = form.cleaned_data['to_date']
			if Project.objects.filter(Q(pro_date__gte=start), Q(pro_date__lte=end)).exists():
				projects = Project.objects.filter(Q(pro_date__gte=start),Q(pro_date__lte=end))

				context_dict ={'AWRI':0,'Murdoch':0,'UQ':0,'UM':0,'UWA':0}

				if Project.nodes.by_node('AWRI'):
					context_dict['AWRI']+=Project.nodes.by_node('AWRI').filter(subtotal__gt=0).count()
				if Project.nodes.by_node('Murdoch'):
					context_dict['Murdoch']+=Project.nodes.by_node('Murdoch').filter(subtotal__gt=0).count()
				if Project.nodes.by_node('UQ'):
					context_dict['UQ']+=Project.nodes.by_node('UQ').filter(subtotal__gt=0).count()
				if Project.nodes.by_node('UM'):
					context_dict['UM']+=Project.nodes.by_node('UM').filter(subtotal__gt=0).count()
				if Project.nodes.by_node('UWA'):
					context_dict['UWA']+=Project.nodes.by_node('UWA').filter(subtotal__gt=0).count()
				
			else:
				context_dict['info'] = "Invalid date input."
			return render(request,'web_MA_DB/invoice_number.html',context_dict)
	else:
		form = 	QuarterForm()

		return render(request,'web_MA_DB/invoice_number.html',{'form':form})


# def usertype_income(request):
# 	projects = Project.objects.all()
# 	data = {'CB':0,'CF':0,'COTH':0,'CP':0,'PCOE_CRC':0,'CSIRO':0,'PDPI':0,
# 	'PMRI':0,'POTH':0,'PRDC':0,'UNI':0}

# 	if projects.filter(usertype='CB'):
# 		data['CB']+=list(projects.filter(usertype='CB').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='CF'):
# 		data['CF']+=list(projects.filter(usertype='CF').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='COTH'):
# 		data['COTH']+=list(projects.filter(usertype='COTH').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='CP'):
# 		data['CP']+=list(projects.filter(usertype='CP').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='PCOE'):
# 		data['PCOE_CRC']+=list(projects.filter(usertype='PCOE').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='CRC'):
# 		data['PCOE_CRC']+=list(projects.filter(usertype='CRC').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='CSIRO'):
# 		data['CSIRO']+=list(projects.filter(usertype='CSIRO').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='PCSIRO'):
# 		data['CSIRO']+=list(projects.filter(usertype='PCSIRO').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='PDPI'):
# 		data['PDPI']+=list(projects.filter(usertype='PDPI').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='PMRI'):
# 		data['PMRI']+=list(projects.filter(usertype='PMRI').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='POTH'):
# 		data['POTH']+=list(projects.filter(usertype='POTH').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='PRDC'):
# 		data['PRDC']+=list(projects.filter(usertype='PRDC').aggregate(Sum('subtotal')).values())[0]
# 	if projects.filter(usertype='UNI'):
# 		data['UNI']+=list(projects.filter(usertype='UNI').aggregate(Sum('subtotal')).values())[0]

# 	return render(request,'web_MA_DB/usertype_income.html',{'CB':data['CB'],'CF':data['CF'],
# 		'CP':data['CP'],'COTH':data['COTH'],'PCOE_CRC':data['PCOE_CRC'],'CSIRO':data['CSIRO'],
# 		'PDPI':data['PDPI'],'PMRI':data['PMRI'],'POTH':data['POTH'],'PRDC':data['PRDC'],'UNI':data['UNI']})

def usertype_income(request):
	context_dict = {}
	if request.method == 'POST':
		form  = QuarterForm(request.POST)
		if form.is_valid():
			start = form.cleaned_data['from_date']
			end  = form.cleaned_data['to_date']

			if Project.objects.filter(Q(pro_date__gte=start), Q(pro_date__lte=end)).exists():
				projects = Project.objects.filter(Q(pro_date__gte=start),Q(pro_date__lte=end))

				
				data = {}
				for t in TYPE_CHOICES:
					data[str(t)]= data.get(str(t),0)+0
					if projects.filter(usertype=t):
						data[str(t)]=data.get(str(t),0)+list(projects.filter(usertype=t).aggregate(Sum('subtotal')).values())[0]

				context_dict['data']= simplejson.dumps(data)
				return render(request,'web_MA_DB/usertype_income.html', context_dict)
			else:
				return render(request,'web_MA_DB/usertype_income.html',{'info':INFO})
	else:
		form = QuarterForm()

	return render(request,'web_MA_DB/usertype_income.html',{'form':form})




def nodes_year_column(request):
	form = FinancialYearForm(request.GET)
	context_dict={}
	projects = Project.objects.all()
	if form.is_valid():
		query_params = form.cleaned_data
		start = int(query_params['from_date'])
		end = int(query_params['to_date'])
		gap = end - start
		years =[]
		context_dict['start']=start
		context_dict['end'] = end
		for i in range(gap):
			years.append(str(start+i)+'-'+str(start+i+1))
		context_dict['years']=simplejson.dumps(years)
		# data = [{'AWRI':0,'Murdoch':0,'UQ':0,'UM':0,'UWA':0} for i in range(gap)]
		awri = []
		murdoch=[]
		uq =[]
		um =[]
		uwa =[]
	
		for i in range(gap):
			tmp_1 = date(start+i,6,30)
			tmp_2 = date(start+i+1,7,1)
			if projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='AWRI'):
				awri.append(list(projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='AWRI').aggregate(Sum('subtotal')).values())[0])
			else: 
				awri.append(0)
			if projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='Murdoch'):
				murdoch.append(list(projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='Murdoch').aggregate(Sum('subtotal')).values())[0])
			else:
				murdoch.append(0)
			if projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='UQ'):
				uq.append(list(projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='UQ').aggregate(Sum('subtotal')).values())[0])
			else:
				uq.append(0)
			if projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='UM'):
				um.append(list(projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='UM').aggregate(Sum('subtotal')).values())[0])
			else:
				um.append(0)
			if projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='UWA'):
				uwa.append(list(projects.filter(Q(pro_date__gte=tmp_1),Q(pro_date__lte=tmp_2),node='UWA').aggregate(Sum('subtotal')).values())[0])
			else:
				uwa.append(0)
		context_dict['AWRI'] = simplejson.dumps(awri)
		context_dict['Murdoch']=simplejson.dumps(murdoch)
		context_dict['UQ']=simplejson.dumps(uq)
		context_dict['UM']=simplejson.dumps(um)
		context_dict['UWA']=simplejson.dumps(uwa)	
	return render(request,'web_MA_DB/nodes_year_column.html',context_dict)

def nodes_year_search(request):
	form = FinancialYearForm()
	return render(request,'web_MA_DB/nodes_year_search.html',{'form':form})
def cat_year_search(request):
	form = QuarterForm()
	return render(request,'web_MA_DB/cat_year_search.html',{'form':form})

def cat_year_column(request):
	form = YearSearchForm(request.GET)
	context_dict={}
	if form. is_valid():
		paras = form.cleaned_data
		start = paras['from_date']
		end = paras['to_date']
		gap = end-start+1
		years =[]
		cat1=[]
		cat2=[]
		cat3=[]

		for i in range(gap):
			years.append(str(start+i))
		context_dict['years']=simplejson.dumps(years)

		projects=Project.objects.all()
		
		for each in years:
			if projects.filter(category='1',pro_date__year=each):
				cat1.append(list(projects.filter(category='1',pro_date__year=each).aggregate(Sum('subtotal')).values())[0])
			else:
				cat1.append(0)
			if projects.filter(category='2',pro_date__year=each):
				cat2.append(list(projects.filter(category='2', pro_date__year=each).aggregate(Sum('subtotal')).values())[0])
			else:
				cat2.append(0)
			if projects.filter(category='3').filter(pro_date__year=each):
				cat3.append(list(projects.filter(category='3', pro_date__year=each).aggregate(Sum('subtotal')).values())[0])
		
		context_dict['cat1']=simplejson.dumps(cat1)
		context_dict['cat2']=simplejson.dumps(cat2)
		context_dict['cat3']=simplejson.dumps(cat3)
	return render(request,'web_MA_DB/cat_year_column.html',context_dict)



def detail_search(request):
	form = SearchForm()
	return render(request,'web_MA_DB/detail_search.html',{'form':form})


def detail_search_results(request):
	context_dict={}
	form = SearchForm(request.GET)
	if form.is_valid():
		query_params = form.cleaned_data
		person = query_params['person']
		start = query_params['from_date']
		if start is None:
			start = date(2003,1,1)
		if query_params['to_date']:
			end = query_params['to_date'] 
		else:
			end = date.today()
		organization = query_params['organization']
		description = query_params['description']
		service = query_params['service']
		instrument = query_params['instrument']

		res = Project.objects.all()

		if query_params['organization']:
			res = res.filter(organization = query_params['organization'])
		if query_params['person']:
			res = res.filter(person__icontains= query_params['person'])
		if query_params['description']:
			res = res.filter(description__icontains =query_params['description'])
		if query_params['service']:
			res = res.filter(service__icontains = query_params['service'])
		if query_params['instrument'] :
			res = res.filter(instrument__icontains = query_params['instrument'])

		res = res.filter(pro_date__gte=start, pro_date__lte=end)


		context_dict['results'] = res
	return render_to_response('web_MA_DB/detail_search_results.html',context_dict)


def MA_search(request):
	form = CustomSearchForm()
	return render(request,'web_MA_DB/MA_search.html',{'form':form})


def MA_search_results(request):
	context_dict={}
	projects = Project.objects.all()
	form = CustomSearchForm(request.GET)
	
	if form.is_valid():
		paras= form.cleaned_data		
		start = paras['from_date']
		end = paras['to_date']
		if 'Category' in paras['choices']:
			context_dict['Category'] = cats(start,end,projects)
		if 'UserDefined_1' in paras['choices']:
			context_dict['UserDefined_1'] = types(start,end,projects)
		if 'UserDefined_2' in paras['choices']:
			context_dict['UserDefined_2'] = fields(start,end,projects) 
		if 'State' in paras['choices']:
			context_dict['State'] = states(start,end,projects)			
		if 'Sum' in paras['choices']:
			context_dict['Sum'] = sum_all(start,end,projects)
		context_dict['from'] = start
		context_dict['to'] = end

		if 'state_excel' in request.POST:	
			all_states = STATES			
			filename = "MA_byState_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			xlsx_data = write_to_excel(filename, context_dict['State'],{'states':all_states})
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response

		if 'type_excel' in request.POST:	
			all_types =[ t.type_id for t in Usertype.objects.filter(~Q(type_id='NULL'))]	
			filename = "MA_byUserDefined1_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			xlsx_data = write_to_excel(filename, context_dict['UserDefined_1'],{'types':types})
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response
		if 'field_excel' in request.POST:
			all_fields = [ f.field_id for f in Userfield.objects.filter(~Q(field_id='NULL'))]
			filename = "MA_byUserDefined2_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			xlsx_data = write_to_excel(filename, context_dict['UserDefined_2'],{'fields':all_fields})
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response
		if 'cat_excel' in request.POST:
			all_cats = ['1','2','3']
			filename = "MA_byCategory_"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			xlsx_data = write_to_excel(filename, context_dict['Category'],{'cats':all_cats})
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response
		
		if 'sum_excel' in request.POST:
			filename = "MA_CustomerDetail"+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)+".xlsx"
			tb_title = "MA "+str(start.year)+"_"+str(start.month)+"-"+str(end.year)+"_"+str(end.month)
			xlsx_data = write_sum_excel(filename,tb_title ,context_dict['Sum'])
			with open(filename,'rb') as fh:
				response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
			return response

		
	else:
		context_dict['info'] ='Invalid date or choices input.'
	
	return render(request,'web_MA_DB/MA_search_results.html',context_dict)



def state_search(request):
	form = StateSearchForm()
	return render(request,'web_MA_DB/state_search.html',{'form':form})


def state_search_result(request):
	form = StateSearchForm(request.GET)
	context_dict = {}

	if form.is_valid():
		query_params = form.cleaned_data
		states = query_params['state']
		start = query_params['from_date']
		end = query_params['to_date']

		data = []

		if Project.objects.filter(pro_date__gte=start,pro_date__lte=end).exists():
			projects = Project.objects.filter(pro_date__gte=start,pro_date__lte=end)
			
			for s in states:
				data.append(s)			
				if projects.filter(state=s).exists():
					data.append(list(projects.filter(state=s).aggregate(Sum('subtotal')).values())[0])
					data.append(list(projects.filter(state=s).aggregate(Sum('cus_count')).values())[0])
					data.append(projects.filter(state=s).count())
				else:
					data.append(0)
					data.append(0)
					data.append(0)

			context_dict['states'] = simplejson.dumps(states)
			context_dict['data'] = [data[x:x+4] for x in range(0, len(data),4)]
			context_dict['from'] = start
			context_dict['to'] = end

			# After aggregation, write results to file to export
			if 'excel' in request.POST:				
				filename = "ByState"+ str(start.year)+"-"+str(end.year)+'_'+str(len(states))+".xlsx"
				xlsx_data = write_state_excel(filename, context_dict)
				with open(filename,'rb') as fh:
					response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
					response['Content-Disposition'] = 'attachment; filename=%s' % filename
				return response
		
		else:
			context_dict['info'] = "Wrong date input."

	return render(request,'web_MA_DB/state_search_result.html',context_dict)



def field_search(request):
	form = FieldForm()
	return render(request,'web_MA_DB/field_search.html',{'form':form})

def field_search_results(request):
	context_dict={}
	form = FieldForm(request.GET)
	if form.is_valid():
		fields = form.cleaned_data['field']
		start = form.cleaned_data['from_date']
		end = form.cleaned_data['to_date']
		data = []

		if Project.objects.filter(pro_date__gte=start,pro_date__lte=end).exists():
			projects = Project.objects.filter(pro_date__gte=start,pro_date__lte=end)

			for f in fields:
				data.append(f)
				if projects.filter(userfield=f).exists():
					data.append(list(projects.filter(userfield=f).aggregate(Sum('subtotal')).values())[0])
					data.append(list(projects.filter(userfield=f).aggregate(Sum('cus_count')).values())[0])
					data.append(projects.filter(userfield=f).count())
				else:
					data.append(0)
					data.append(0)
					data.append(0)
			context_dict['fields'] = simplejson.dumps(fields)
			context_dict['data'] = [data[x:x+4] for x in range(0, len(data),4)]
			context_dict['from'] = start
			context_dict['to'] = end

			if 'excel' in request.POST:				
				filename = "ByUserDefined_2__"+ str(start.year)+"-"+str(end.year)+".xlsx"
				xlsx_data = write_state_excel(filename, context_dict)
				with open(filename,'rb') as fh:
					response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
					response['Content-Disposition'] = 'attachment; filename=%s' % filename
				return response

		else:
			context_dict['info'] = "Wrong date input."

	return render(request,'web_MA_DB/field_search_results.html',context_dict)


def type_search(request):
	form = TypeForm()
	return render(request,'web_MA_DB/type_search.html',{'form':form})

def type_search_results(request):
	context_dict={}
	form = TypeForm(request.GET)
	if form.is_valid():
		types = form.cleaned_data['usertype']
		start = form.cleaned_data['from_date']
		end = form.cleaned_data['to_date']
		data = []

		if Project.objects.filter(pro_date__gte=start,pro_date__lte=end).exists():
			projects = Project.objects.filter(pro_date__gte=start,pro_date__lte=end)

			for t in types:
				data.append(t)
				if projects.filter(usertype=t).exists():
					data.append(list(projects.filter(usertype=t).aggregate(Sum('subtotal')).values())[0])
					data.append(list(projects.filter(usertype=t).aggregate(Sum('cus_count')).values())[0])
					data.append(projects.filter(usertype=t).count())
				else:
					data.append(0)
					data.append(0)
					data.append(0)
			context_dict['types'] = simplejson.dumps(types)
			context_dict['data'] = [data[x:x+4] for x in range(0, len(data),4)]
			context_dict['from'] = start
			context_dict['to'] = end

			if 'excel' in request.POST:				
				filename = "ByUserDefined_1__"+ str(start.year)+"-"+str(end.year)+".xlsx"
				xlsx_data = write_state_excel(filename, context_dict)
				with open(filename,'rb') as fh:
					response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
					response['Content-Disposition'] = 'attachment; filename=%s' % filename
				return response


		else:
			context_dict['info'] = "Wrong date input."

	return render(request,'web_MA_DB/type_search_results.html',context_dict)

def int_ext_search(request):
	form = IntExtForm()
	return render(request,'web_MA_DB/int_ext_search.html',{'form':form})

def int_ext_results(request):
	context_dict={}
	form = IntExtForm(request.GET)
	if form.is_valid():
		intexts = form.cleaned_data['intext']
		start = form.cleaned_data['from_date']
		end = form.cleaned_data['to_date']
		data = []

		if Project.objects.filter(pro_date__gte=start,pro_date__lte=end).exists():
			projects = Project.objects.filter(pro_date__gte=start,pro_date__lte=end)

			for i in intexts:
				data.append(i)
				if i == 'Internal':
					if projects.filter(int_ext='INT').exists():
						data.append(list(projects.filter(int_ext='INT').aggregate(Sum('subtotal')).values())[0])
						data.append(list(projects.filter(int_ext='INT').aggregate(Sum('cus_count')).values())[0])
						data.append(projects.filter(int_ext='INT').count())
					else:
						data.append(0)
						data.append(0)
						data.append(0)
				if i == 'External':
					if projects.filter(int_ext='EXT').exists():
						data.append(list(projects.filter(int_ext='EXT').aggregate(Sum('subtotal')).values())[0])
						data.append(list(projects.filter(int_ext='EXT').aggregate(Sum('cus_count')).values())[0])
						data.append(projects.filter(int_ext='EXT').count())
					else:
						data.append(0)
						data.append(0)
						data.append(0)
			context_dict['intexts'] = simplejson.dumps(intexts)
			context_dict['data'] = [data[x:x+4] for x in range(0, len(data),4)]
			context_dict['from'] = start
			context_dict['to'] = end

			if 'excel' in request.POST:				
				filename = "ByInt_Ext_"+ str(start.year)+"-"+str(end.year)+".xlsx"
				xlsx_data = write_state_excel(filename, context_dict)
				with open(filename,'rb') as fh:
					response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
					response['Content-Disposition'] = 'attachment; filename=%s' % filename
				return response


		else:
			context_dict['info'] = "Wrong date input."

	return render(request,'web_MA_DB/int_ext_results.html',context_dict)

def node_search(request):
	form = InvoiceSearchForm()
	return render(request,'web_MA_DB/node_search.html',{'form':form})

def node_search_results(request):

	# def grouped(l,n):
	# 	for i in range(0,len(l),n):
	# 		yield l[i:i+n]

	context_dict={}
	form = InvoiceSearchForm(request.GET)

	if form.is_valid():
		query_params = form.cleaned_data
		nodes = query_params['node']
		start = form.cleaned_data['from_date']
		end = form.cleaned_data['to_date']
		context_dict['nodes']=simplejson.dumps(nodes)
		data=[]
				
		if Project.objects.filter(pro_date__gte=start,pro_date__lte=end).exists():
			projects = Project.objects.filter(pro_date__gte=start,pro_date__lte=end)

			for n in nodes:
				data.append(n)
				if projects.filter(node=n).exists():
					data.append(list(projects.filter(node=n).aggregate(Sum('subtotal')).values())[0])
					data.append(list(projects.filter(node=n).aggregate(Sum('cus_count')).values())[0])
					data.append(projects.filter(node=n).count())
				else:
					data.append(0)
					data.append(0)
					data.append(0)
			context_dict['nodes'] = simplejson.dumps(nodes)
			context_dict['data'] = [data[x:x+4] for x in range(0, len(data),4)]
			context_dict['from'] = start
			context_dict['to'] = end

			if 'excel' in request.POST:				
				filename = "ByNode_"+ str(start.year)+"-"+str(end.year)+".xlsx"
				xlsx_data = write_state_excel(filename, context_dict)
				with open(filename,'rb') as fh:
					response = HttpResponse(fh.read(),content_type ='application/vnd.ms-excel')			
					response['Content-Disposition'] = 'attachment; filename=%s' % filename
				return response

		else:
			context_dict['info'] = "Wrong date input."

	return render(request,'web_MA_DB/node_search_results.html',context_dict)



def um_reconcil_search(request):
	form = ReconcilliationForm()
	return render(request,'web_MA_DB/um_reconcil_search.html',{'form':form})

def um_reconcil_results(request):
	form = ReconcilliationForm(request.GET)
	context_dict={}
	if form.is_valid():
		start = form.cleaned_data['from_date']
		end = form.cleaned_data['to_date']
		data = list(Invoice.objects.filter(inv_date__gte=start, inv_date__lte = end).aggregate(Sum('subtotal')).values())[0]
	context_dict['start'] = start
	context_dict['end'] = end
	context_dict['data'] = data

	return render(request,'web_MA_DB/um_reconcil_results.html',context_dict)


def readProjectSheet(request):

	states = ['VIC','QLD','NSW','SA','ACT','NZ','TAS','WA','NT']

	if request.method == 'POST':
		form = UploadFileForm(request.POST, request.FILES)
		if form.is_valid():
			sheetname = form.cleaned_data['sheet_name']
			start = form.cleaned_data['from_row']
			end = form.cleaned_data['to_row']
			newfile = ProjectSheet(file_name=request.FILES['file_name'],sheet_name=sheetname)
			newfile.save()

			book = openpyxl.load_workbook(newfile.file_name,data_only = True,read_only=True)
			all_sheets= book.get_sheet_names()
			if sheetname not in all_sheets:
				return render(request,"web_MA_DB/upload.html",{"warning":"Wrong file or non-exist sheet."})
			sheet = book.get_sheet_by_name(sheetname)
			database = MySQLdb.connect(host="localhost",user = "root",passwd = "WYBZDhehe0221", db="MA_DATABASE")
			cursor = database.cursor()

			for row in range(start,end+1):
				node = sheet['A'+str(row)].value
				pro_date = sheet['B'+str(row)].value
				description = sheet['C'+str(row)].value
				service = sheet['D'+str(row)].value
				instrument = sheet['E'+str(row)].value
				person = sheet['F'+str(row)].value
				organization = sheet['G'+str(row)].value
				num_sample = sheet['H'+str(row)].internal_value
				category = sheet['I'+str(row)].value
				int_ext = sheet['J'+str(row)].value
				state=sheet['K'+str(row)].value
				country = sheet['L'+str(row)].value
				usertype = sheet['M'+str(row)].value
				userfield = sheet['N'+str(row)].value
				subtotal = sheet['O'+ str(row)].value
				cus_count = sheet['P'+str(row)].value

				if isinstance(pro_date,date):
					pro_date = pro_date.isoformat()
				else:
					pro_date = date.today().isoformat()
				
				if person is None:
					person = 'Null'
				
				if organization is None:
					organization = "Null"

				if int_ext is not None:
					int_ext = int_ext.upper()
				
				if state is None:
					state = 'Overseas'
				else:
					state= state.upper()
					if state not in set(states):
						state ='Overseas'
				if cus_count is None:
					cus_count ="0"
				
				if usertype is not None:
					usertype = usertype.upper()
				else:
					usertype = "NULL"
				
				if userfield is not None:
					userfield=userfield.upper()
				else:
					userfield = "NULL"
				
				if not isinstance(num_sample,int):
					num_sample=0
				
				if subtotal is None:
					subtotal =0.00
				else:
					subtotal=decimal.Decimal("%.2f" % subtotal)


				if not Usertype.objects.filter(type_id=usertype).exists():
					tmp_query = """ INSERT INTO Usertype (type_id,type_name) VALUES (%s,%s)"""
					tmp_values = (usertype,"Null") 
					cursor.execute(tmp_query,tmp_values)
					database.commit()
				if not Userfield.objects.filter(field_id=userfield).exists():
					pre_query = """ INSERT INTO Userfield (field_id,field_name) VALUES (%s,%s)"""
					pre_values = (userfield,"Null") 
					cursor.execute(pre_query,pre_values)
					database.commit()

				query = """INSERT INTO Project (node, pro_date,description,service,instrument,person,organization,num_sample,category,int_ext,state,country,usertype,userfield,subtotal,cus_count) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
				values=(node, pro_date,description,service,instrument,person,
					organization,num_sample,category,int_ext,state,country,
					usertype,userfield,subtotal,cus_count)
				cursor.execute(query,values)
				database.commit()

			# End of for loop			
			cursor.close()
			database.close()
			
			return render(request,'web_MA_DB/upload.html',{"info":"Successful upload!"})
	else:
		form = UploadFileForm()
		return render(request, 'web_MA_DB/upload.html', {'form': form})
	

def readInvoiceSheet(request):

	if request.method == 'POST':
		form = UploadFileForm(request.POST, request.FILES)
		if form.is_valid():
			sheetname = form.cleaned_data['sheet_name']
			start = form.cleaned_data['from_row']
			end = form.cleaned_data['to_row']
			newfile = InvoiceSheet(file_name=request.FILES['file_name'],sheet_name=sheetname)
			newfile.save()

			book = openpyxl.load_workbook(newfile.file_name,data_only = True,read_only=True)
			sheet = book.get_sheet_by_name(sheetname)
			database = MySQLdb.connect(host="localhost",user = "root",passwd = "WYBZDhehe0221", db="MA_DATABASE")
			cursor = database.cursor()

			for row in range(start,end+1):
				inv_date = sheet['A'+str(row)].value
				inv_no = sheet['B'+str(row)].value
				quote_no = sheet['D'+str(row)].value
				ma_staff = sheet['E'+str(row)].value
				description = sheet['F'+str(row)].value
				service = sheet['G'+str(row)].value
				instrument = sheet['H'+str(row)].value
				person = sheet['I'+str(row)].value
				address = sheet['J'+str(row)].value
				num_sample = sheet['K'+str(row)].internal_value
				category = sheet['L'+str(row)].value
				int_ext = sheet['M'+str(row)].value
				state=sheet['N'+str(row)].value
				country = sheet['O'+str(row)].value
				usertype = sheet['P'+str(row)].value
				userfield = sheet['Q'+str(row)].value
				subtotal = sheet['R'+ str(row)].value

				if isinstance(inv_date,date):
					inv_date = inv_date.isoformat()
				else:
					inv_date = date.today().isoformat()
				if person is None:
					person = 'Null'
				if ma_staff is None:
					ma_staff = "Null"
				if address is None:
					address = "Null"
				if usertype is not None:
					usertype = usertype.upper()
				else:
					usertype = "NULL"
				if userfield is not None:
					userfield=userfield.upper()
				else:
					userfield = "NULL"
				if not isinstance(num_sample,int):
					num_sample=0
				if subtotal is None:
					subtotal =0.00
				else:
					subtotal=decimal.Decimal("%.2f" % subtotal)

				# field_check = userfield.split(';')
				# if len(field_check)==1:
				
				if not Usertype.objects.filter(type_id=usertype).exists():
					tmp_query = """ INSERT INTO Usertype (type_id,type_name) VALUES (%s,%s)"""
					tmp_values = (usertype,"Null") 
					cursor.execute(tmp_query,tmp_values)
					database.commit()
				if not Userfield.objects.filter(field_id=userfield).exists():
					pre_query = """ INSERT INTO Userfield (field_id,field_name) VALUES (%s,%s)"""
					pre_values = (userfield,"Null") 
					cursor.execute(pre_query,pre_values)
					database.commit()

				query = """INSERT INTO Invoice (inv_date,inv_no,quote_no,MA_staff,description,service,instrument,person,address,num_sample,category,int_ext,state,country,usertype,userfield,subtotal) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
				values=(inv_date,inv_no,quote_no, ma_staff,description,service,instrument,person,
					address,num_sample,category,int_ext,state,country,
					usertype,userfield,subtotal)
				cursor.execute(query,values)
				database.commit()
				
			# End of for loop			
			cursor.close()
			database.close()
			return HttpResponseRedirect('/web_MA_DB/homepage/')

	else:
		form = UploadFileForm()
	return render(request, 'web_MA_DB/import_invoice.html', {'form': form})


def readQuoteSheet(request):
	if request.method == 'POST':
		form = UploadFileForm(request.POST, request.FILES)
		if form.is_valid():
			sheetname = form.cleaned_data['sheet_name']
			newfile = QuoteSheet(file_name=request.FILES['file_name'])
			newfile.save()

			book = openpyxl.load_workbook(newfile.file_name,data_only = True,read_only=True)
			sheet = book.get_sheet_by_name(sheetname)

			database = MySQLdb.connect(host="localhost",user = "root",passwd = "WYBZDhehe0221", db="MA_DATABASE")
			cursor = database.cursor()


			for row in range(3,sheet.max_row-1):
				qt_name = sheet['A'+str(row)].value
				qt_year = sheet['B'+str(row)].value
				version = sheet['C'+str(row)].value
				client = sheet['E'+str(row)].value
				company = sheet['F'+str(row)].value
				ma_staff = sheet['G'+str(row)].value
				qt_date = sheet['H'+str(row)].value
				for_grant = sheet['I'+str(row)].value
				accepted = sheet['J'+str(row)].value
				invoiced = sheet['K'+str(row)].value
				comment = sheet['L'+str(row)].value

				if version is None:
					version = 'v1'
				else:
					if version.startswith('_'):
						version = version.strip('_')
					else:
						version = version

				if client is None:
					client = 'Null'
				if company is None:
					company= 'Null'
				if ma_staff is None:
					ma_staff = 'Null'
				if isinstance(qt_date,date):
					qt_date = qt_date.isoformat()
				else:
					qt_date = date.today().isoformat()
				if for_grant is not None:
					for_grant = '1'
				if for_grant is None and accepted is not None:
					if 'grant' in accepted:
						for_grant='1'
					else:
						for_grant = '0'
				if accepted is not None:
					accepted = '1'
				else:
					accepted = '0'
				if invoiced is not None:
					invoiced = '1'
				else:
					invoiced = '0'
				query = """INSERT INTO Quote(qt_name,qt_year,version,client,company,ma_staff,qt_date,for_grant,accepted,invoiced,comment) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
				values =(qt_name,qt_year,version,client,company,ma_staff,qt_date,for_grant,accepted,invoiced,comment)
				cursor.execute(query,values)
				database.commit()

			cursor.close()
			database.close()
			return HttpResponseRedirect('/web_MA_DB/homepage/')

	else:
		form = UploadFileForm()
	return render(request, 'web_MA_DB/import_quotes.html', {'form': form})


def update_field(request):
	context_dict = {}
	if request.method =='POST':
		form = UpdateFieldForm(request.POST)
		if form.is_valid():
			field_id = form.cleaned_data['field_id']
			field_name = form.cleaned_data['field_name']

			if Userfield.objects.filter(field_id=field_id).exists():
				Userfield.objects.filter(field_id=field_id).update(field_name=field_name)
			else:
				f = Userfield.objects.create(field_id=field_id,field_name=field_name)
				f.save()
			
			new_fields = Userfield.objects.filter(~Q(field_id="NULL"))
			context_dict['new_fields'] = new_fields
			return render(request,'web_MA_DB/update_field.html',context_dict)
			
	else:
		form = UpdateFieldForm()
		context_dict['form']=form
	return render(request,'web_MA_DB/update_field.html',context_dict)


def update_type(request):
	context_dict = {}
	if request.method =='POST':
		form = UpdateFieldForm(request.POST)
		if form.is_valid():
			type_id = form.cleaned_data['field_id']
			type_name = form.cleaned_data['field_name']

			if Usertype.objects.filter(type_id=type_id).exists():
				Usertype.objects.filter(type_id=type_id).update(type_name=type_name)
			else:
				f = Usertype.objects.create(type_id=type_id,type_name=type_name)
				f.save()
			new_types = Usertype.objects.filter(~Q(type_id="NULL"))
			context_dict['new_types'] = new_types
			return render(request,'web_MA_DB/update_type.html',context_dict)
	else:
		form = UpdateFieldForm()
		context_dict['form']=form
		return render(request,'web_MA_DB/update_type.html',context_dict)



def register(request):
	registered =False

	if request.method =='POST':
		user_form = UserForm(data = request.POST)

		if user_form.is_valid():
			user=user_form.save(commit=False)
			user.set_password(user.password)
			user.save()
			registered = True
		else:
			print(user_form.errors)
	else:
		user_form = UserForm()

	return render(request,'web_MA_DB/register.html',{'form':user_form,'registered':registered})


def admin_login(request):

	if request.method=='POST':
		username = request.POST.get('username')
		password = request.POST.get('password')

		user = authenticate(username = username,password=password)

		if user is not None:
			# if user.is_active():
			login(request, user)
			return HttpResponseRedirect('/web_MA_DB/homepage/')
		# 	else:
		# 		return HttpResponse('Your administrator account is disabled.')
		else:
			messages.warning(request, 'Invalid username or password.')
			return HttpResponseRedirect('/web_MA_DB/login/')
	else:
		return render(request,'web_MA_DB/login.html',{})


def admin_logout(request):
	
	logout(request)
	return HttpResponseRedirect('/web_MA_DB/login/')












